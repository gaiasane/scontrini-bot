import os
import re
import json
from datetime import datetime, timedelta
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import pytesseract
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import asyncio

# Configurazione Tesseract


# Token del bot
TOKEN = '8537568170:AAEJHhQIBbR9jPErgah47yjJOOiN67Lkdcg'

# Nomi delle persone
PERSONE = ['Gaia', 'Ari']

# Categorie complete
CATEGORIE = [
    '🍕 Cibo e bevande',
    '🛒 Spesa',
    '🚗 Trasporti',
    '🏠 Casa e utenze',
    '💊 Salute e benessere',
    '🎭 Divertimento',
    '👕 Shopping',
    '🎓 Istruzione',
    '✈️ Viaggi',
    '💡 Bollette e utenze',
    '📱 Telefono e internet',
    '🔧 Manutenzione casa',
    '💰 Altro'
]

# Stati della conversazione
PAGATORE, CATEGORIA, DESCRIZIONE, IMPORTO = range(4)
# Stati per spesa manuale
MANUALE_PAGATORE, MANUALE_CATEGORIA, MANUALE_DESCRIZIONE, MANUALE_IMPORTO = range(4, 8)
# Stati per spese ricorrenti
RIC_DESCRIZIONE, RIC_IMPORTO, RIC_PAGATORE, RIC_CATEGORIA, RIC_GIORNO, RIC_FINE = range(8, 14)
# Stati per modifica
MOD_CAMPO, MOD_VALORE = range(14, 16)

# Percorsi file
EXCEL_FILE = 'spese_condivise.xlsx'
SCONTRINI_FOLDER = 'scontrini'
USERS_FILE = 'users.json'
RICORRENTI_FILE = 'ricorrenti.json'

# Crea cartella scontrini se non esiste
if not os.path.exists(SCONTRINI_FOLDER):
    os.makedirs(SCONTRINI_FOLDER)

def salva_user_id(user_id, username):
    """Salva l'ID utente per le notifiche"""
    users = {}
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            users = json.load(f)
    
    users[str(user_id)] = username
    
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f)

def carica_users():
    """Carica gli ID utenti salvati"""
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    return {}

def salva_ricorrente(ricorrente):
    """Salva una spesa ricorrente"""
    ricorrenti = carica_ricorrenti()
    ricorrenti.append(ricorrente)
    
    with open(RICORRENTI_FILE, 'w') as f:
        json.dump(ricorrenti, f, indent=2)

def carica_ricorrenti():
    """Carica le spese ricorrenti"""
    if os.path.exists(RICORRENTI_FILE):
        with open(RICORRENTI_FILE, 'r') as f:
            return json.load(f)
    return []

def elimina_ricorrente(indice):
    """Elimina una spesa ricorrente"""
    ricorrenti = carica_ricorrenti()
    if 0 <= indice < len(ricorrenti):
        del ricorrenti[indice]
        with open(RICORRENTI_FILE, 'w') as f:
            json.dump(ricorrenti, f, indent=2)
        return True
    return False

def ottieni_nome_foglio(data=None):
    """Restituisce il nome del foglio per un dato mese"""
    if data is None:
        data = datetime.now()
    return data.strftime("%Y-%m")

def crea_foglio_mese(wb, nome_foglio, saldo_precedente=0):
    """Crea un nuovo foglio per il mese"""
    if nome_foglio in wb.sheetnames:
        return wb[nome_foglio]
    
    ws = wb.create_sheet(nome_foglio)
    
    # Intestazione saldo precedente
    ws['A1'] = 'SALDO MESE PRECEDENTE'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    ws['B1'] = f'€{saldo_precedente:.2f}'
    ws['B1'].font = Font(bold=True, size=12)
    ws['B1'].alignment = Alignment(horizontal='right')
    
    # Intestazioni colonne
    headers = ['Data', 'Chi ha pagato', 'Categoria', 'Descrizione', 'Importo €', 'Quota Gaia', 'Quota Ari', 'File Scontrino']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Imposta larghezza colonne
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 30
    
    wb.save(EXCEL_FILE)
    return ws

def crea_excel():
    """Crea un nuovo file Excel se non esiste"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        # Rimuovi il foglio di default
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Crea il primo foglio per il mese corrente
        nome_foglio = ottieni_nome_foglio()
        crea_foglio_mese(wb, nome_foglio)
        
        wb.save(EXCEL_FILE)

def calcola_saldo_mese(wb, nome_foglio):
    """Calcola il saldo di un mese specifico"""
    if nome_foglio not in wb.sheetnames:
        return 0
    
    ws = wb[nome_foglio]
    
    totale_gaia = 0
    totale_ari = 0
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[4]:  # Se c'è un importo
            pagatore = row[1]
            importo = float(row[4])
            
            if pagatore == 'Gaia':
                totale_gaia += importo
            elif pagatore == 'Ari':
                totale_ari += importo
    
    # Saldo: positivo se Gaia ha pagato di più, negativo se Ari ha pagato di più
    return (totale_gaia - totale_ari) / 2

def ottieni_saldo_precedente(wb, data):
    """Ottiene il saldo del mese precedente"""
    mese_precedente = (data.replace(day=1) - timedelta(days=1))
    nome_foglio_prec = ottieni_nome_foglio(mese_precedente)
    
    return calcola_saldo_mese(wb, nome_foglio_prec)

def aggiungi_spesa(data, pagatore, categoria, descrizione, importo, nome_file=''):
    """Aggiunge una spesa al file Excel"""
    crea_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # Ottieni o crea il foglio del mese
    data_obj = datetime.strptime(data, "%d/%m/%Y")
    nome_foglio = ottieni_nome_foglio(data_obj)
    
    if nome_foglio not in wb.sheetnames:
        saldo_prec = ottieni_saldo_precedente(wb, data_obj)
        ws = crea_foglio_mese(wb, nome_foglio, saldo_prec)
    else:
        ws = wb[nome_foglio]
    
    # Calcola le quote (diviso a metà)
    quota = importo / 2
    
    # Aggiungi riga
    nuova_riga = [
        data,
        pagatore,
        categoria,
        descrizione,
        importo,
        quota,
        quota,
        nome_file
    ]
    
    ws.append(nuova_riga)
    wb.save(EXCEL_FILE)

def ottieni_ultime_spese(n=10):
    """Ottiene le ultime N spese"""
    if not os.path.exists(EXCEL_FILE):
        return []
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nome_foglio = ottieni_nome_foglio()
    
    if nome_foglio not in wb.sheetnames:
        wb.close()
        return []
    
    ws = wb[nome_foglio]
    
    spese = []
    for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if row[4]:  # Se c'è un importo
            spese.append({
                'riga': idx,
                'data': row[0],
                'pagatore': row[1],
                'categoria': row[2],
                'descrizione': row[3],
                'importo': float(row[4])
            })
    
    wb.close()
    
    # Ritorna le ultime N
    return spese[-n:]

def elimina_spesa_da_excel(riga):
    """Elimina una spesa dal file Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nome_foglio = ottieni_nome_foglio()
    
    if nome_foglio not in wb.sheetnames:
        wb.close()
        return False
    
    ws = wb[nome_foglio]
    ws.delete_rows(riga)
    wb.save(EXCEL_FILE)
    wb.close()
    return True

def modifica_spesa_excel(riga, campo, nuovo_valore):
    """Modifica un campo di una spesa"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nome_foglio = ottieni_nome_foglio()
    
    if nome_foglio not in wb.sheetnames:
        wb.close()
        return False
    
    ws = wb[nome_foglio]
    
    # Mappa dei campi
    campi_map = {
        'pagatore': 2,  # Colonna B
        'categoria': 3,  # Colonna C
        'descrizione': 4,  # Colonna D
        'importo': 5  # Colonna E
    }
    
    if campo not in campi_map:
        wb.close()
        return False
    
    col = campi_map[campo]
    ws.cell(row=riga, column=col).value = nuovo_valore
    
    # Se modifichiamo l'importo, ricalcola le quote
    if campo == 'importo':
        quota = float(nuovo_valore) / 2
        ws.cell(row=riga, column=6).value = quota  # Quota Gaia
        ws.cell(row=riga, column=7).value = quota  # Quota Ari
    
    wb.save(EXCEL_FILE)
    wb.close()
    return True

def processa_spese_ricorrenti():
    """Aggiunge le spese ricorrenti del mese se non ancora aggiunte"""
    ricorrenti = carica_ricorrenti()
    oggi = datetime.now()
    
    for ric in ricorrenti:
        # Controlla se la spesa è ancora valida
        if ric['tipo_fine'] == 'data':
            data_fine = datetime.strptime(ric['data_fine'], "%d/%m/%Y")
            if oggi > data_fine:
                continue
        
        # Controlla se questo mese è già stata aggiunta
        giorno = int(ric['giorno'])
        if oggi.day < giorno:
            continue
        
        data_spesa = oggi.replace(day=giorno).strftime("%d/%m/%Y")
        
        # Verifica se già esiste nel foglio
        wb = openpyxl.load_workbook(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else None
        if wb:
            nome_foglio = ottieni_nome_foglio(oggi)
            if nome_foglio in wb.sheetnames:
                ws = wb[nome_foglio]
                
                # Cerca se la spesa ricorrente è già presente questo mese
                trovata = False
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if row[3] == f"[AUTO] {ric['descrizione']}" and row[0] == data_spesa:
                        trovata = True
                        break
                
                if trovata:
                    continue
            wb.close()
        
        # Aggiungi la spesa ricorrente
        aggiungi_spesa(
            data_spesa,
            ric['pagatore'],
            ric['categoria'],
            f"[AUTO] {ric['descrizione']}",
            float(ric['importo']),
            'Spesa ricorrente'
        )

def estrai_importo_da_immagine(image_path):
    """Estrae l'importo da un'immagine usando OCR"""
    return None

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /start"""
    # Salva l'ID utente per le notifiche
    user_id = update.effective_user.id
    username = update.effective_user.first_name or "User"
    salva_user_id(user_id, username)
    
    await update.message.reply_text(
        "👋 Ciao! Sono il bot delle spese condivise!\n\n"
        "📸 Inviami una foto dello scontrino per registrare una spesa\n\n"
        "📝 Comandi disponibili:\n"
        "/start - Mostra questo messaggio\n"
        "/aggiungi - Aggiungi spesa senza foto\n"
        "/report - Report del mese corrente\n"
        "/elimina - Elimina una spesa\n"
        "/modifica - Modifica una spesa\n"
        "/ricorrente - Aggiungi spesa ricorrente\n"
        "/lista_ricorrenti - Vedi spese ricorrenti\n"
        "/elimina_ricorrente - Elimina spesa ricorrente\n"
        "/cancella - Annulla operazione"
    )

# ========== GESTIONE SPESE CON FOTO ==========

async def ricevi_scontrino(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve lo scontrino e inizia la conversazione"""
    photo = update.message.photo[-1]
    
    # Salva l'immagine
    file = await context.bot.get_file(photo.file_id)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_file = f"scontrino_{timestamp}.jpg"
    percorso = os.path.join(SCONTRINI_FOLDER, nome_file)
    await file.download_to_drive(percorso)
    
    # Salva il percorso nel contesto
    context.user_data['scontrino'] = percorso
    context.user_data['nome_file'] = nome_file
    
    # Estrai importo
    await update.message.reply_text("🔍 Sto analizzando lo scontrino...")
    importo = estrai_importo_da_immagine(percorso)
    
    if importo:
        context.user_data['importo_ocr'] = importo
        await update.message.reply_text(f"✅ Scontrino ricevuto! Ho trovato: €{importo:.2f}")
    else:
        await update.message.reply_text("✅ Scontrino ricevuto! (Non sono riuscito a leggere l'importo)")
    
    # Chiedi chi ha pagato
    keyboard = [[persona] for persona in PERSONE]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text("Chi ha pagato?", reply_markup=reply_markup)
    
    return PAGATORE

async def ricevi_pagatore(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve chi ha pagato"""
    pagatore = update.message.text
    
    if pagatore not in PERSONE:
        await update.message.reply_text("❌ Seleziona un nome valido!")
        keyboard = [[persona] for persona in PERSONE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Chi ha pagato?", reply_markup=reply_markup)
        return PAGATORE
    
    context.user_data['pagatore'] = pagatore
    
    # Chiedi la categoria
    keyboard = [[cat] for cat in CATEGORIE]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text("Che categoria?", reply_markup=reply_markup)
    
    return CATEGORIA

async def ricevi_categoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve la categoria"""
    categoria = update.message.text
    
    if categoria not in CATEGORIE:
        await update.message.reply_text("❌ Seleziona una categoria valida!")
        keyboard = [[cat] for cat in CATEGORIE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Che categoria?", reply_markup=reply_markup)
        return CATEGORIA
    
    context.user_data['categoria'] = categoria
    
    # Chiedi la descrizione
    await update.message.reply_text(
        "📝 Inserisci una descrizione:\n(es: 'Spesa Conad', 'Cena pizzeria', 'Bolletta Enel')",
        reply_markup=ReplyKeyboardRemove()
    )
    
    return DESCRIZIONE

async def ricevi_descrizione(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve la descrizione"""
    descrizione = update.message.text
    context.user_data['descrizione'] = descrizione
    
    # Se abbiamo già l'importo dall'OCR, chiedi conferma
    if 'importo_ocr' in context.user_data:
        importo = context.user_data['importo_ocr']
        keyboard = [['✅ Conferma', '✏️ Modifica']]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        
        await update.message.reply_text(
            f"💰 Importo rilevato: €{importo:.2f}\nConfermi?",
            reply_markup=reply_markup
        )
        return IMPORTO
    else:
        await update.message.reply_text(
            "💰 Quanto hai speso?\n(es: 12.50 o 12,50)",
            reply_markup=ReplyKeyboardRemove()
        )
        return IMPORTO

async def ricevi_importo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve l'importo finale"""
    risposta = update.message.text
    
    # Se conferma l'importo OCR
    if risposta == '✅ Conferma' and 'importo_ocr' in context.user_data:
        importo = context.user_data['importo_ocr']
    elif risposta == '✏️ Modifica':
        await update.message.reply_text(
            "💰 Inserisci l'importo corretto:\n(es: 12.50 o 12,50)",
            reply_markup=ReplyKeyboardRemove()
        )
        return IMPORTO
    else:
        # Altrimenti inserimento manuale
        try:
            importo = float(risposta.replace(',', '.').replace('€', '').strip())
        except:
            await update.message.reply_text("❌ Importo non valido! Inserisci un numero (es: 12.50)")
            return IMPORTO
    
    # Salva tutto
    data = datetime.now().strftime("%d/%m/%Y")
    pagatore = context.user_data['pagatore']
    categoria = context.user_data['categoria']
    descrizione = context.user_data['descrizione']
    nome_file = context.user_data['nome_file']
    
    aggiungi_spesa(data, pagatore, categoria, descrizione, importo, nome_file)
    
    quota = importo / 2
    
    await update.message.reply_text(
        f"✅ Spesa registrata!\n\n"
        f"📅 Data: {data}\n"
        f"💳 Chi ha pagato: {pagatore}\n"
        f"📂 Categoria: {categoria}\n"
        f"📝 Descrizione: {descrizione}\n"
        f"💰 Totale: €{importo:.2f}\n"
        f"📊 Quota Gaia: €{quota:.2f}\n"
        f"📊 Quota Ari: €{quota:.2f}",
        reply_markup=ReplyKeyboardRemove()
    )
    
    # Pulisci i dati
    context.user_data.clear()
    
    return ConversationHandler.END

# ========== GESTIONE SPESE MANUALI (SENZA FOTO) ==========

async def aggiungi_manuale(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inizia l'aggiunta di una spesa senza foto"""
    keyboard = [[persona] for persona in PERSONE]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text("Chi ha pagato?", reply_markup=reply_markup)
    
    return MANUALE_PAGATORE

async def ricevi_pagatore_manuale(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve chi ha pagato (manuale)"""
    pagatore = update.message.text
    
    if pagatore not in PERSONE:
        await update.message.reply_text("❌ Seleziona un nome valido!")
        keyboard = [[persona] for persona in PERSONE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Chi ha pagato?", reply_markup=reply_markup)
        return MANUALE_PAGATORE
    
    context.user_data['pagatore'] = pagatore
    
    # Chiedi la categoria
    keyboard = [[cat] for cat in CATEGORIE]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text("Che categoria?", reply_markup=reply_markup)
    
    return MANUALE_CATEGORIA

async def ricevi_categoria_manuale(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve la categoria (manuale)"""
    categoria = update.message.text
    
    if categoria not in CATEGORIE:
        await update.message.reply_text("❌ Seleziona una categoria valida!")
        keyboard = [[cat] for cat in CATEGORIE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Che categoria?", reply_markup=reply_markup)
        return MANUALE_CATEGORIA
    
    context.user_data['categoria'] = categoria
    
    # Chiedi la descrizione
    await update.message.reply_text(
        "📝 Inserisci una descrizione:\n(es: 'Spesa Conad', 'Cena pizzeria', 'Bolletta Enel')",
        reply_markup=ReplyKeyboardRemove()
    )
    
    return MANUALE_DESCRIZIONE

async def ricevi_descrizione_manuale(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve la descrizione (manuale)"""
    descrizione = update.message.text
    context.user_data['descrizione'] = descrizione
    
    await update.message.reply_text(
        "💰 Quanto hai speso?\n(es: 12.50 o 12,50)",
        reply_markup=ReplyKeyboardRemove()
    )
    
    return MANUALE_IMPORTO

async def ricevi_importo_manuale(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve l'importo (manuale)"""
    try:
        importo = float(update.message.text.replace(',', '.').replace('€', '').strip())
    except:
        await update.message.reply_text("❌ Importo non valido! Inserisci un numero (es: 12.50)")
        return MANUALE_IMPORTO
    
    # Salva tutto
    data = datetime.now().strftime("%d/%m/%Y")
    pagatore = context.user_data['pagatore']
    categoria = context.user_data['categoria']
    descrizione = context.user_data['descrizione']
    
    aggiungi_spesa(data, pagatore, categoria, descrizione, importo, 'Inserimento manuale')
    
    quota = importo / 2
    
    await update.message.reply_text(
        f"✅ Spesa registrata!\n\n"
        f"📅 Data: {data}\n"
        f"💳 Chi ha pagato: {pagatore}\n"
        f"📂 Categoria: {categoria}\n"
        f"📝 Descrizione: {descrizione}\n"
        f"💰 Totale: €{importo:.2f}\n"
        f"📊 Quota Gaia: €{quota:.2f}\n"
        f"📊 Quota Ari: €{quota:.2f}",
        reply_markup=ReplyKeyboardRemove()
    )
    
    # Pulisci i dati
    context.user_data.clear()
    
    return ConversationHandler.END

# ========== ELIMINAZIONE SPESE ==========

async def elimina_spesa_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra le ultime spese per eliminarle"""
    spese = ottieni_ultime_spese(10)
    
    if not spese:
        await update.message.reply_text("📋 Non ci sono spese da eliminare!")
        return ConversationHandler.END
    
    msg = "🗑️ ULTIME 10 SPESE:\n\n"
    
    for i, spesa in enumerate(spese, 1):
        msg += f"{i}. {spesa['data']} - €{spesa['importo']:.2f}\n"
        msg += f"   {spesa['pagatore']} - {spesa['categoria']}\n"
        msg += f"   {spesa['descrizione']}\n\n"
    
    msg += "Quale spesa vuoi eliminare? (Scrivi il numero)"
    
    context.user_data['spese_da_eliminare'] = spese
    
    await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
    
    return 0

async def elimina_spesa_conferma(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Elimina la spesa selezionata"""
    try:
        scelta = int(update.message.text)
        spese = context.user_data.get('spese_da_eliminare', [])
        
        if 1 <= scelta <= len(spese):
            spesa = spese[scelta - 1]
            
            if elimina_spesa_da_excel(spesa['riga']):
                await update.message.reply_text(
                    f"✅ Spesa eliminata:\n"
                    f"📅 {spesa['data']}\n"
                    f"💰 €{spesa['importo']:.2f}\n"
                    f"📝 {spesa['descrizione']}"
                )
            else:
                await update.message.reply_text("❌ Errore durante l'eliminazione!")
        else:
            await update.message.reply_text("❌ Numero non valido!")
    except:
        await update.message.reply_text("❌ Inserisci un numero valido!")
    
    context.user_data.clear()
    return ConversationHandler.END

# ========== MODIFICA SPESE ==========

async def modifica_spesa_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra le ultime spese per modificarle"""
    spese = ottieni_ultime_spese(10)
    
    if not spese:
        await update.message.reply_text("📋 Non ci sono spese da modificare!")
        return ConversationHandler.END
    
    msg = "✏️ ULTIME 10 SPESE:\n\n"
    
    for i, spesa in enumerate(spese, 1):
        msg += f"{i}. {spesa['data']} - €{spesa['importo']:.2f}\n"
        msg += f"   {spesa['pagatore']} - {spesa['categoria']}\n"
        msg += f"   {spesa['descrizione']}\n\n"
    
    msg += "Quale spesa vuoi modificare? (Scrivi il numero)"
    
    context.user_data['spese_da_modificare'] = spese
    
    await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
    
    return MOD_CAMPO

async def modifica_spesa_campo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Chiede quale campo modificare"""
    try:
        scelta = int(update.message.text)
        spese = context.user_data.get('spese_da_modificare', [])
        
        if 1 <= scelta <= len(spese):
            spesa = spese[scelta - 1]
            context.user_data['spesa_da_modificare'] = spesa
            
            keyboard = [
                ['💳 Chi ha pagato'],
                ['📂 Categoria'],
                ['📝 Descrizione'],
                ['💰 Importo']
            ]
            reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
            
            await update.message.reply_text(
                f"Cosa vuoi modificare di questa spesa?\n\n"
                f"📅 {spesa['data']}\n"
                f"💳 {spesa['pagatore']}\n"
                f"📂 {spesa['categoria']}\n"
                f"📝 {spesa['descrizione']}\n"
                f"💰 €{spesa['importo']:.2f}",
                reply_markup=reply_markup
            )
            
            return MOD_VALORE
        else:
            await update.message.reply_text("❌ Numero non valido!")
            context.user_data.clear()
            return ConversationHandler.END
    except:
        await update.message.reply_text("❌ Inserisci un numero valido!")
        context.user_data.clear()
        return ConversationHandler.END

async def modifica_spesa_valore(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve il campo da modificare e chiede il nuovo valore"""
    scelta = update.message.text
    
    campo_map = {
        '💳 Chi ha pagato': 'pagatore',
        '📂 Categoria': 'categoria',
        '📝 Descrizione': 'descrizione',
        '💰 Importo': 'importo'
    }
    
    if scelta not in campo_map:
        await update.message.reply_text("❌ Selezione non valida!")
        context.user_data.clear()
        return ConversationHandler.END
    
    campo = campo_map[scelta]
    context.user_data['campo_da_modificare'] = campo
    
    if campo == 'pagatore':
        keyboard = [[persona] for persona in PERSONE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Chi ha pagato?", reply_markup=reply_markup)
    elif campo == 'categoria':
        keyboard = [[cat] for cat in CATEGORIE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Nuova categoria?", reply_markup=reply_markup)
    elif campo == 'descrizione':
        await update.message.reply_text("Nuova descrizione?", reply_markup=ReplyKeyboardRemove())
    elif campo == 'importo':
        await update.message.reply_text("Nuovo importo? (es: 12.50)", reply_markup=ReplyKeyboardRemove())
    
    return 99  # Stato per ricevere il nuovo valore

async def modifica_spesa_salva(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Salva la modifica"""
    nuovo_valore = update.message.text
    campo = context.user_data['campo_da_modificare']
    spesa = context.user_data['spesa_da_modificare']
    
    # Validazione
    if campo == 'pagatore' and nuovo_valore not in PERSONE:
        await update.message.reply_text("❌ Nome non valido!")
        context.user_data.clear()
        return ConversationHandler.END
    
    if campo == 'categoria' and nuovo_valore not in CATEGORIE:
        await update.message.reply_text("❌ Categoria non valida!")
        context.user_data.clear()
        return ConversationHandler.END
    
    if campo == 'importo':
        try:
            nuovo_valore = float(nuovo_valore.replace(',', '.').replace('€', '').strip())
        except:
            await update.message.reply_text("❌ Importo non valido!")
            context.user_data.clear()
            return ConversationHandler.END
    
    # Modifica nel file Excel
    if modifica_spesa_excel(spesa['riga'], campo, nuovo_valore):
        await update.message.reply_text(
            f"✅ Spesa modificata!\n\n"
            f"Campo: {campo}\n"
            f"Nuovo valore: {nuovo_valore}",
            reply_markup=ReplyKeyboardRemove()
        )
    else:
        await update.message.reply_text("❌ Errore durante la modifica!")
    
    context.user_data.clear()
    return ConversationHandler.END

# ========== GESTIONE SPESE RICORRENTI ==========

async def aggiungi_ricorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inizia l'aggiunta di una spesa ricorrente"""
    await update.message.reply_text(
        "📝 Inserisci la descrizione della spesa ricorrente:\n"
        "(es: 'Affitto', 'Netflix', 'Abbonamento palestra')",
        reply_markup=ReplyKeyboardRemove()
    )
    return RIC_DESCRIZIONE

async def ricevi_descrizione_ricorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve la descrizione della spesa ricorrente"""
    context.user_data['ric_descrizione'] = update.message.text
    
    await update.message.reply_text("💰 Qual è l'importo mensile?\n(es: 50.00)")
    return RIC_IMPORTO

async def ricevi_importo_ricorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve l'importo della spesa ricorrente"""
    try:
        importo = float(update.message.text.replace(',', '.').replace('€', '').strip())
        context.user_data['ric_importo'] = importo
    except:
        await update.message.reply_text("❌ Importo non valido! Inserisci un numero (es: 50.00)")
        return RIC_IMPORTO
    
    keyboard = [[persona] for persona in PERSONE]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("💳 Chi paga?", reply_markup=reply_markup)
    return RIC_PAGATORE

async def ricevi_pagatore_ricorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve chi paga la spesa ricorrente"""
    pagatore = update.message.text
    
    if pagatore not in PERSONE:
        await update.message.reply_text("❌ Seleziona un nome valido!")
        keyboard = [[persona] for persona in PERSONE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Chi paga?", reply_markup=reply_markup)
        return RIC_PAGATORE
    
    context.user_data['ric_pagatore'] = pagatore
    
    keyboard = [[cat] for cat in CATEGORIE]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("📂 Che categoria?", reply_markup=reply_markup)
    return RIC_CATEGORIA

async def ricevi_categoria_ricorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve la categoria della spesa ricorrente"""
    categoria = update.message.text
    
    if categoria not in CATEGORIE:
        await update.message.reply_text("❌ Seleziona una categoria valida!")
        keyboard = [[cat] for cat in CATEGORIE]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Che categoria?", reply_markup=reply_markup)
        return RIC_CATEGORIA
    
    context.user_data['ric_categoria'] = categoria
    
    await update.message.reply_text(
        "📅 Che giorno del mese viene addebitata?\n"
        "(Inserisci un numero da 1 a 28)\n"
        "(es: 1 per il primo del mese, 15 per il quindici)",
        reply_markup=ReplyKeyboardRemove()
    )
    return RIC_GIORNO

async def ricevi_giorno_ricorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve il giorno del mese"""
    try:
        giorno = int(update.message.text)
        if giorno < 1 or giorno > 28:
            raise ValueError()
        context.user_data['ric_giorno'] = giorno
    except:
        await update.message.reply_text("❌ Inserisci un numero valido da 1 a 28!")
        return RIC_GIORNO
    
    keyboard = [
        ['♾️ Mai (continua sempre)'],
        ['📅 Data specifica'],
        ['✋ Manuale (elimino io)']
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("🏁 Quando finisce questa spesa ricorrente?", reply_markup=reply_markup)
    return RIC_FINE

async def ricevi_fine_ricorrente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Riceve quando finisce la spesa ricorrente"""
    risposta = update.message.text
    
    if risposta == '♾️ Mai (continua sempre)':
        tipo_fine = 'mai'
        data_fine = None
    elif risposta == '✋ Manuale (elimino io)':
        tipo_fine = 'manuale'
        data_fine = None
    elif risposta == '📅 Data specifica':
        await update.message.reply_text(
            "📅 Inserisci la data di fine:\n(formato: GG/MM/AAAA, es: 31/12/2025)",
            reply_markup=ReplyKeyboardRemove()
        )
        context.user_data['attendi_data'] = True
        return RIC_FINE
    elif 'attendi_data' in context.user_data:
        # Validazione data
        try:
            data_fine = datetime.strptime(risposta, "%d/%m/%Y").strftime("%d/%m/%Y")
            tipo_fine = 'data'
            del context.user_data['attendi_data']
        except:
            await update.message.reply_text("❌ Data non valida! Usa il formato GG/MM/AAAA (es: 31/12/2025)")
            return RIC_FINE
    else:
        await update.message.reply_text("❌ Seleziona un'opzione valida!")
        keyboard = [
            ['♾️ Mai (continua sempre)'],
            ['📅 Data specifica'],
            ['✋ Manuale (elimino io)']
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("Quando finisce questa spesa ricorrente?", reply_markup=reply_markup)
        return RIC_FINE
    
    # Salva la spesa ricorrente
    ricorrente = {
        'descrizione': context.user_data['ric_descrizione'],
        'importo': context.user_data['ric_importo'],
        'pagatore': context.user_data['ric_pagatore'],
        'categoria': context.user_data['ric_categoria'],
        'giorno': context.user_data['ric_giorno'],
        'tipo_fine': tipo_fine,
        'data_fine': data_fine
    }
    
    salva_ricorrente(ricorrente)
    
    # Messaggio di conferma
    msg = (
        f"✅ Spesa ricorrente aggiunta!\n\n"
        f"📝 {ricorrente['descrizione']}\n"
        f"💰 €{ricorrente['importo']:.2f}\n"
        f"💳 Paga: {ricorrente['pagatore']}\n"
        f"📂 {ricorrente['categoria']}\n"
        f"📅 Giorno: {ricorrente['giorno']} di ogni mese\n"
    )
    
    if tipo_fine == 'mai':
        msg += "🏁 Continua all'infinito"
    elif tipo_fine == 'manuale':
        msg += "🏁 Finisce quando la elimini manualmente"
    else:
        msg += f"🏁 Finisce il: {data_fine}"
    
    await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
    
    context.user_data.clear()
    return ConversationHandler.END

async def lista_ricorrenti(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra tutte le spese ricorrenti"""
    ricorrenti = carica_ricorrenti()
    
    if not ricorrenti:
        await update.message.reply_text("📋 Non hai ancora spese ricorrenti!")
        return
    
    msg = "📋 SPESE RICORRENTI:\n\n"
    
    for i, ric in enumerate(ricorrenti, 1):
        msg += f"{i}. {ric['descrizione']}\n"
        msg += f"   💰 €{ric['importo']:.2f} - {ric['pagatore']}\n"
        msg += f"   📂 {ric['categoria']}\n"
        msg += f"   📅 Giorno {ric['giorno']}\n"
        
        if ric['tipo_fine'] == 'mai':
            msg += "   🏁 Continua sempre\n"
        elif ric['tipo_fine'] == 'manuale':
            msg += "   🏁 Manuale\n"
        else:
            msg += f"   🏁 Fine: {ric['data_fine']}\n"
        
        msg += "\n"
    
    await update.message.reply_text(msg)

async def elimina_ricorrente_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inizia il processo di eliminazione di una spesa ricorrente"""
    ricorrenti = carica_ricorrenti()
    
    if not ricorrenti:
        await update.message.reply_text("📋 Non hai spese ricorrenti da eliminare!")
        return ConversationHandler.END
    
    msg = "🗑️ Quale spesa ricorrente vuoi eliminare?\n\n"
    
    for i, ric in enumerate(ricorrenti, 1):
        msg += f"{i}. {ric['descrizione']} - €{ric['importo']:.2f}\n"
    
    msg += "\nInserisci il numero della spesa da eliminare:"
    
    await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
    
    return 0  # Stato per ricevere il numero

async def elimina_ricorrente_conferma(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Elimina la spesa ricorrente selezionata"""
    try:
        indice = int(update.message.text) - 1
        ricorrenti = carica_ricorrenti()
        
        if 0 <= indice < len(ricorrenti):
            ric_eliminata = ricorrenti[indice]
            elimina_ricorrente(indice)
            
            await update.message.reply_text(
                f"✅ Spesa ricorrente eliminata:\n"
                f"📝 {ric_eliminata['descrizione']}\n"
                f"💰 €{ric_eliminata['importo']:.2f}"
            )
        else:
            await update.message.reply_text("❌ Numero non valido!")
    except:
        await update.message.reply_text("❌ Inserisci un numero valido!")
    
    return ConversationHandler.END

# ========== REPORT ==========

async def genera_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Genera un report mensile dettagliato"""
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("📊 Nessuna spesa registrata ancora!")
        return
    
    # Prima processa le spese ricorrenti
    processa_spese_ricorrenti()
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nome_foglio = ottieni_nome_foglio()
    
    if nome_foglio not in wb.sheetnames:
        await update.message.reply_text("📊 Nessuna spesa per questo mese!")
        wb.close()
        return
    
    ws = wb[nome_foglio]
    
    # Leggi saldo precedente
    saldo_precedente = 0
    if ws['B1'].value:
        try:
            saldo_str = str(ws['B1'].value).replace('€', '').strip()
            saldo_precedente = float(saldo_str)
        except:
            pass
    
    # Calcola totali
    totale_generale = 0
    totale_per_categoria = {}
    spese_per_categoria = {}
    totale_gaia_pagato = 0
    totale_ari_pagato = 0
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[4]:  # Se c'è un importo
            data = row[0]
            pagatore = row[1]
            categoria = row[2]
            descrizione = row[3]
            importo = float(row[4])
            
            totale_generale += importo
            
            if pagatore == 'Gaia':
                totale_gaia_pagato += importo
            elif pagatore == 'Ari':
                totale_ari_pagato += importo
            
            # Totale per categoria
            if categoria not in totale_per_categoria:
                totale_per_categoria[categoria] = 0
                spese_per_categoria[categoria] = []
            
            totale_per_categoria[categoria] += importo
            spese_per_categoria[categoria].append({
                'data': data,
                'importo': importo,
                'descrizione': descrizione
            })
    
    wb.close()
    
    # Crea messaggio report
    mese_corrente = datetime.now().strftime("%B %Y")
    
    messaggio = f"📊 REPORT {mese_corrente.upper()}\n"
    messaggio += "=" * 30 + "\n\n"
    
    # Saldo precedente
    if saldo_precedente != 0:
        messaggio += f"💼 Saldo mese precedente: €{saldo_precedente:.2f}\n"
        if saldo_precedente > 0:
            messaggio += "   (Ari doveva a Gaia)\n"
        else:
            messaggio += "   (Gaia doveva ad Ari)\n"
        messaggio += "\n"
    
    messaggio += f"💰 TOTALE SPESO: €{totale_generale:.2f}\n\n"
    
    # Dettaglio per categoria
    messaggio += "📂 DETTAGLIO PER CATEGORIA:\n"
    messaggio += "-" * 30 + "\n"
    
    for cat in sorted(totale_per_categoria.keys(), key=lambda x: totale_per_categoria[x], reverse=True):
        tot = totale_per_categoria[cat]
        percentuale = (tot / totale_generale * 100) if totale_generale > 0 else 0
        
        messaggio += f"\n{cat}\n"
        messaggio += f"Totale: €{tot:.2f} ({percentuale:.1f}%)\n"
        
        # Elenca le singole spese
        for spesa in spese_per_categoria[cat]:
            messaggio += f"  • {spesa['data']} - €{spesa['importo']:.2f} - {spesa['descrizione']}\n"
    
    messaggio += "\n" + "=" * 30 + "\n"
    messaggio += f"💳 CHI HA PAGATO:\n"
    messaggio += f"Gaia: €{totale_gaia_pagato:.2f}\n"
    messaggio += f"Ari: €{totale_ari_pagato:.2f}\n\n"
    
    # Calcola saldo
    quota_gaia = totale_generale / 2
    quota_ari = totale_generale / 2
    
    saldo_gaia = totale_gaia_pagato - quota_gaia
    saldo_ari = totale_ari_pagato - quota_ari
    
    # Includi saldo precedente
    saldo_totale = saldo_gaia + saldo_precedente
    
    messaggio += "💸 SALDO:\n"
    
    if abs(saldo_totale) < 0.01:
        messaggio += "✅ Siete in pari!"
    elif saldo_totale > 0:
        messaggio += f"Ari deve a Gaia: €{abs(saldo_totale):.2f}"
    else:
        messaggio += f"Gaia deve ad Ari: €{abs(saldo_totale):.2f}"
    
    # Invia il messaggio (diviso se troppo lungo)
    if len(messaggio) > 4000:
        parti = [messaggio[i:i+4000] for i in range(0, len(messaggio), 4000)]
        for parte in parti:
            await update.message.reply_text(parte)
    else:
        await update.message.reply_text(messaggio)
    
    # Invia anche il file Excel
    await update.message.reply_document(
        document=open(EXCEL_FILE, 'rb'),
        filename=f'Spese_{nome_foglio}.xlsx'
    )

async def invia_report_automatico(application):
    """Invia il report automaticamente il 1° del mese"""
    users = carica_users()
    
    if not users:
        return
    
    # Genera il report del mese precedente
    mese_precedente = (datetime.now().replace(day=1) - timedelta(days=1))
    nome_foglio_prec = ottieni_nome_foglio(mese_precedente)
    
    if not os.path.exists(EXCEL_FILE):
        return
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    if nome_foglio_prec not in wb.sheetnames:
        wb.close()
        return
    
    ws = wb[nome_foglio_prec]
    
    # Calcola totali del mese precedente
    totale_generale = 0
    totale_per_categoria = {}
    spese_per_categoria = {}
    totale_gaia_pagato = 0
    totale_ari_pagato = 0
    saldo_precedente = 0
    
    # Leggi saldo precedente
    if ws['B1'].value:
        try:
            saldo_str = str(ws['B1'].value).replace('€', '').strip()
            saldo_precedente = float(saldo_str)
        except:
            pass
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[4]:
            data = row[0]
            pagatore = row[1]
            categoria = row[2]
            descrizione = row[3]
            importo = float(row[4])
            
            totale_generale += importo
            
            if pagatore == 'Gaia':
                totale_gaia_pagato += importo
            elif pagatore == 'Ari':
                totale_ari_pagato += importo
            
            if categoria not in totale_per_categoria:
                totale_per_categoria[categoria] = 0
                spese_per_categoria[categoria] = []
            
            totale_per_categoria[categoria] += importo
            spese_per_categoria[categoria].append({
                'data': data,
                'importo': importo,
                'descrizione': descrizione
            })
    
    wb.close()
    
    # Crea messaggio
    mese_nome = mese_precedente.strftime("%B %Y")
    
    messaggio = f"📊 REPORT AUTOMATICO - {mese_nome.upper()}\n"
    messaggio += "=" * 30 + "\n\n"
    
    if saldo_precedente != 0:
        messaggio += f"💼 Saldo precedente: €{saldo_precedente:.2f}\n\n"
    
    messaggio += f"💰 TOTALE SPESO: €{totale_generale:.2f}\n\n"
    
    messaggio += "📂 PER CATEGORIA:\n"
    messaggio += "-" * 30 + "\n"
    
    for cat in sorted(totale_per_categoria.keys(), key=lambda x: totale_per_categoria[x], reverse=True):
        tot = totale_per_categoria[cat]
        percentuale = (tot / totale_generale * 100) if totale_generale > 0 else 0
        
        messaggio += f"\n{cat}\n"
        messaggio += f"Totale: €{tot:.2f} ({percentuale:.1f}%)\n"
        
        for spesa in spese_per_categoria[cat]:
            messaggio += f"  • {spesa['data']} - €{spesa['importo']:.2f} - {spesa['descrizione']}\n"
    
    messaggio += "\n" + "=" * 30 + "\n"
    messaggio += f"💳 CHI HA PAGATO:\n"
    messaggio += f"Gaia: €{totale_gaia_pagato:.2f}\n"
    messaggio += f"Ari: €{totale_ari_pagato:.2f}\n\n"
    
    quota_gaia = totale_generale / 2
    saldo_gaia = totale_gaia_pagato - quota_gaia
    saldo_totale = saldo_gaia + saldo_precedente
    
    messaggio += "💸 SALDO:\n"
    
    if abs(saldo_totale) < 0.01:
        messaggio += "✅ Siete in pari!"
    elif saldo_totale > 0:
        messaggio += f"Ari deve a Gaia: €{abs(saldo_totale):.2f}"
    else:
        messaggio += f"Gaia deve ad Ari: €{abs(saldo_totale):.2f}"
    
    # Invia a tutti gli utenti registrati
    for user_id in users.keys():
        try:
            if len(messaggio) > 4000:
                parti = [messaggio[i:i+4000] for i in range(0, len(messaggio), 4000)]
                for parte in parti:
                    await application.bot.send_message(chat_id=int(user_id), text=parte)
            else:
                await application.bot.send_message(chat_id=int(user_id), text=messaggio)
            
            # Invia file Excel
            await application.bot.send_document(
                chat_id=int(user_id),
                document=open(EXCEL_FILE, 'rb'),
                filename=f'Spese_{nome_foglio_prec}.xlsx'
            )
        except Exception as e:
            print(f"Errore invio a {user_id}: {e}")

async def controlla_report_mensile(application):
    """Controlla se è il 1° del mese e invia il report"""
    while True:
        ora_corrente = datetime.now()
        
        # Se è il 1° del mese alle 9:00
        if ora_corrente.day == 1 and ora_corrente.hour == 9 and ora_corrente.minute == 0:
            await invia_report_automatico(application)
            # Aspetta 1 ora per evitare invii multipli
            await asyncio.sleep(3600)
        
        # Controlla ogni 60 secondi
        await asyncio.sleep(60)

async def cancella(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cancella l'operazione"""
    context.user_data.clear()
    await update.message.reply_text("❌ Operazione annullata.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

def main():
    """Avvia il bot"""
    # Crea l'applicazione
    application = Application.builder().token(TOKEN).build()
    
    # Conversation handler per foto
    conv_handler_foto = ConversationHandler(
        entry_points=[MessageHandler(filters.PHOTO, ricevi_scontrino)],
        states={
            PAGATORE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_pagatore)],
            CATEGORIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_categoria)],
            DESCRIZIONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_descrizione)],
            IMPORTO: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_importo)],
        },
        fallbacks=[CommandHandler('cancella', cancella)],
    )
    
    # Conversation handler per spesa manuale
    conv_handler_manuale = ConversationHandler(
        entry_points=[CommandHandler('aggiungi', aggiungi_manuale)],
        states={
            MANUALE_PAGATORE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_pagatore_manuale)],
            MANUALE_CATEGORIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_categoria_manuale)],
            MANUALE_DESCRIZIONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_descrizione_manuale)],
            MANUALE_IMPORTO: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_importo_manuale)],
        },
        fallbacks=[CommandHandler('cancella', cancella)],
    )
    
    # Conversation handler per eliminazione spese
    conv_handler_elimina = ConversationHandler(
        entry_points=[CommandHandler('elimina', elimina_spesa_start)],
        states={
            0: [MessageHandler(filters.TEXT & ~filters.COMMAND, elimina_spesa_conferma)],
        },
        fallbacks=[CommandHandler('cancella', cancella)],
    )
    
    # Conversation handler per modifica spese
    conv_handler_modifica = ConversationHandler(
        entry_points=[CommandHandler('modifica', modifica_spesa_start)],
        states={
            MOD_CAMPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, modifica_spesa_campo)],
            MOD_VALORE: [MessageHandler(filters.TEXT & ~filters.COMMAND, modifica_spesa_valore)],
            99: [MessageHandler(filters.TEXT & ~filters.COMMAND, modifica_spesa_salva)],
        },
        fallbacks=[CommandHandler('cancella', cancella)],
    )
    
    # Conversation handler per spese ricorrenti
    conv_handler_ricorrente = ConversationHandler(
        entry_points=[CommandHandler('ricorrente', aggiungi_ricorrente)],
        states={
            RIC_DESCRIZIONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_descrizione_ricorrente)],
            RIC_IMPORTO: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_importo_ricorrente)],
            RIC_PAGATORE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_pagatore_ricorrente)],
            RIC_CATEGORIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_categoria_ricorrente)],
            RIC_GIORNO: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_giorno_ricorrente)],
            RIC_FINE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ricevi_fine_ricorrente)],
        },
        fallbacks=[CommandHandler('cancella', cancella)],
    )
    
    # Conversation handler per eliminazione ricorrenti
    conv_handler_elimina_ric = ConversationHandler(
        entry_points=[CommandHandler('elimina_ricorrente', elimina_ricorrente_start)],
        states={
            0: [MessageHandler(filters.TEXT & ~filters.COMMAND, elimina_ricorrente_conferma)],
        },
        fallbacks=[CommandHandler('cancella', cancella)],
    )
    
    # Aggiungi handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("report", genera_report))
    application.add_handler(CommandHandler("lista_ricorrenti", lista_ricorrenti))
    application.add_handler(conv_handler_foto)
    application.add_handler(conv_handler_manuale)
    application.add_handler(conv_handler_elimina)
    application.add_handler(conv_handler_modifica)
    application.add_handler(conv_handler_ricorrente)
    application.add_handler(conv_handler_elimina_ric)
    
    # Avvia il task per il controllo mensile
    application.job_queue.run_repeating(
        lambda context: asyncio.create_task(controlla_report_mensile(context.application)),
        interval=60,
        first=0
    )
    
    # Avvia il bot
    print("🤖 Bot avviato! Premi Ctrl+C per fermare.")
    print("📱 Cerca il bot su Telegram e invia /start")
    application.run_polling()

if __name__ == '__main__':
    main()